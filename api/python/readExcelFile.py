#!/usr/bin/env python3
import json
import sys
from datetime import datetime, date, time as time_type
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)


def to_date_str(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%d.%m.%Y") if hasattr(val, 'strftime') else str(val)
    return str(val).strip()


def to_time_str(val):
    if val is None:
        return "00:00"
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time_type):
        return f"{val.hour:02d}:{val.minute:02d}"
    if isinstance(val, float):
        total = round(val * 24 * 60)
        return f"{total // 60:02d}:{total % 60:02d}"
    return str(val).strip()


def to_duration_str(val):
    if val is None:
        return None
    if isinstance(val, float):
        total = round(val * 24 * 60)
        return f"{total // 60}:{total % 60:02d}"
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    if isinstance(val, time_type):
        return f"{val.hour}:{val.minute:02d}"
    return str(val).strip()


def build_iso(date_val, time_val):
    d = to_date_str(date_val)
    t = to_time_str(time_val)
    if not d:
        return None
    for fmt in ("%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(f"{d} {t}", fmt).isoformat()
        except ValueError:
            continue
    return None


def parse_sheet(sheet):
    driver_name = sheet.cell(row=1, column=2).value or "Unknown"

    records = []
    for row in sheet.iter_rows(min_row=11, values_only=True):
        if not any(row):
            continue

        date_s, time_s, date_e, time_e, dolzina, aktivnost, reg, posadka = (
            row[0], row[1], row[2], row[3], row[4], row[5],
            row[6] if len(row) > 6 else None,
            row[7] if len(row) > 7 else None,
        )

        if not aktivnost:
            continue

        zacetek = build_iso(date_s, time_s)
        konec   = build_iso(date_e, time_e)
        if not zacetek or not konec:
            continue

        records.append({
            "voznik":     str(driver_name),
            "zacetek":    zacetek,
            "konec":      konec,
            "dolzina":    to_duration_str(dolzina),
            "aktivnost":  str(aktivnost),
            "registerska": str(reg) if reg else None,
            "posadka":    "Da" if str(posadka).strip().lower() in ("da", "yes", "true", "1") else "Ne",
        })

    return str(driver_name), records


def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    drivers = []
    for sheet in wb.worksheets:
        name, records = parse_sheet(sheet)
        if records:
            drivers.append({"voznik": name, "records": records})
    return drivers


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "File path required"}))
        sys.exit(1)

    file_path = sys.argv[1]
    if not Path(file_path).exists():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        sys.exit(1)

    try:
        drivers = parse_excel(file_path)
        if not drivers:
            print(json.dumps({"error": "No data found in file"}))
            sys.exit(1)

        result = {
            "voznik":  drivers[0]["voznik"],
            "records": drivers[0]["records"],
            "drivers": drivers,
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(0)

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
