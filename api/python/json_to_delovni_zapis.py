#!/usr/bin/env python3
"""
Convert JSON driver activity log to Excel delovni zapis (work record).
Usage: python json_to_delovni_zapis.py input.json output.xlsx
"""

import json
import sys
import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Slovenska imena dni ──────────────────────────────────────────────────────
SL_DAYS = ['ponedeljek','torek','sreda','četrtek','petek','sobota','nedelja']

# ── Konstante za tipe aktivnosti ─────────────────────────────────────────────
ACT_VOZNJA         = 'Vožnja'
ACT_RAZPOLOZLJIVOST = 'Razpoložljivost'
ACT_POCITEK        = 'Počitek'
ACT_DELO           = 'Delo'
ACT_NEZNANO        = 'Neznano'

# ── Preslikava aktivnosti → stolpec ──────────────────────────────────────────
ACTIVITY_COL = {
    ACT_VOZNJA:          5,   # E
    ACT_DELO:            6,   # F
    ACT_RAZPOLOZLJIVOST: 17,  # Q
}

STANJE_MAP = {
    'POCITEK':         ACT_POCITEK,
    'VOZNJA':          ACT_VOZNJA,
    'DELO':            ACT_DELO,
    'RAZPOLOZLJIVOST': ACT_RAZPOLOZLJIVOST,
    'NEZNANO':         ACT_NEZNANO,
}

# ── Konfiguracija nočnega dela ───────────────────────────────────────────────
NIGHT_START = 22
NIGHT_END   = 6

HH_MM_SS  = '[HH]:mm:ss'
GREY_FILL = PatternFill('solid', start_color='CCCCCC', end_color='CCCCCC')

COL_WIDTHS = {
    'A': 45.7, 'B': 24.7, 'C': 22.7, 'D': 45.7, 'E': 15.7,
    'F': 13.0, 'G': 23.7, 'H': 17.7, 'I': 21.7, 'J': 15.7,
    'K': 20.7, 'L': 15.7, 'M': 23.7, 'N': 17.7, 'O': 23.7,
    'P': 21.7, 'Q': 18.7, 'R': 16.7, 'S': 20.7, 'T': 30.7,
    'U': 24.7, 'V': 15.7,
}

HEADERS = [
    'Datum', 'Začetek delovnega dne', 'Konec delovnega dne', 'Prisotnost',
    ACT_VOZNJA, ACT_DELO, 'Zakonski delovni čas', 'Druga dela (M)',
    'Skupni delovni čas', 'Nočno delo', 'Koriščenje ur (M)',
    'Dopust', ACT_RAZPOLOZLJIVOST, 'Izraba odmora', 'Odmori in počitki',
    'Nalaganje/razkladanje', 'Posadka',
]


def td(minutes: float) -> datetime.timedelta:
    return datetime.timedelta(seconds=int(minutes * 60))


def parse_hhmm(s: str) -> float:
    if not s:
        return 0.0
    parts = s.split(':')
    return int(parts[0]) * 60 + int(parts[1])


def load_json(path: str):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def group_by_driver_date(records):
    data        = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    raw_records = defaultdict(lambda: defaultdict(list))
    all_drivers = []

    for driver_obj in records:
        uporabnik = driver_obj.get('uporabnik', {})
        driver    = f"{uporabnik.get('ime', '')} {uporabnik.get('priimek', '')}".strip()
        all_drivers.append(driver)

        for rec in driver_obj.get('voznjeMesec', []):
            stanje = rec.get('stanje', '')
            act    = STANJE_MAP.get(stanje, stanje.capitalize() if stanje else '')
            dur    = rec.get('trajanje_min', 0)

            zacetek_str = rec.get('zacetek', '')
            if zacetek_str:
                zacetek_dt = datetime.datetime.fromisoformat(zacetek_str.replace('Z', '+00:00'))
                date = zacetek_dt.date()
            else:
                date = datetime.date.today()

            data[driver][date][act]           += dur
            raw_records[driver][date].append(rec)

    return data, all_drivers, raw_records


def all_dates_in_range(date_set):
    if not date_set:
        return []
    mn, mx = min(date_set), max(date_set)
    days, cur = [], mn
    while cur <= mx:
        days.append(cur)
        cur += datetime.timedelta(days=1)
    return days


def parse_iso_datetime(dt_str: str) -> datetime.datetime:
    if not dt_str:
        return None
    try:
        return datetime.datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
    except ValueError:
        return None


def get_hour_minute(dt: datetime.datetime) -> tuple:
    if not dt:
        return None
    return (dt.hour, dt.minute)


def is_night_work(zacetek_dt: datetime.datetime, konc_dt: datetime.datetime,
                  night_start=NIGHT_START, night_end=NIGHT_END) -> float:
    if not zacetek_dt or not konc_dt:
        return 0.0

    tz            = zacetek_dt.tzinfo
    total_overlap = 0.0
    check_date    = zacetek_dt.date() - datetime.timedelta(days=1)
    end_date      = konc_dt.date()

    while check_date <= end_date:
        next_day     = check_date + datetime.timedelta(days=1)
        night_begin  = datetime.datetime(check_date.year, check_date.month, check_date.day,
                                         night_start, 0, tzinfo=tz)
        night_finish = datetime.datetime(next_day.year, next_day.month, next_day.day,
                                         night_end, 0, tzinfo=tz)

        overlap_start = max(zacetek_dt, night_begin)
        overlap_end   = min(konc_dt, night_finish)
        if overlap_end > overlap_start:
            total_overlap += (overlap_end - overlap_start).total_seconds() / 60

        check_date += datetime.timedelta(days=1)

    return total_overlap


def _build_day_times(day_records):
    """Izračuna start_time, end_time in nocno_delo_mins iz surovih zapisov dneva."""
    start_time      = None
    end_time        = None
    nocno_delo_mins = 0.0

    for rec in day_records:
        act_type   = STANJE_MAP.get(rec.get('stanje', ''), '')
        if act_type not in (ACT_VOZNJA, ACT_DELO):
            continue

        zacetek_dt = parse_iso_datetime(rec.get('zacetek', ''))
        konc_dt    = parse_iso_datetime(rec.get('konec', ''))

        if zacetek_dt and (start_time is None or zacetek_dt < start_time):
            start_time = zacetek_dt
        if konc_dt and (end_time is None or konc_dt > end_time):
            end_time = konc_dt

        nocno_delo_mins += is_night_work(zacetek_dt, konc_dt)

    return start_time, end_time, nocno_delo_mins


def _write_day_row(ws, current_row, d, acts, day_records, week_totals, total_mins):
    """Zapiše eno podatkovno vrstico za dan."""
    thin_border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    day_label = f'{d.strftime("%d.%m.%Y")} {SL_DAYS[d.weekday()]}'
    cell_day  = ws.cell(row=current_row, column=1, value=day_label)
    cell_day.alignment = center_align

    start_time, end_time, nocno_delo_mins = _build_day_times(day_records)

    voznja_mins  = acts.get(ACT_VOZNJA, 0.0)
    delo_mins    = acts.get(ACT_DELO, 0.0)
    razpol_mins  = acts.get(ACT_RAZPOLOZLJIVOST, 0.0)
    rest_mins    = acts.get(ACT_POCITEK, 0.0)
    skupni_mins  = voznja_mins + delo_mins + razpol_mins

    if start_time:
        c = ws.cell(row=current_row, column=2, value=start_time.strftime('%H:%M'))
        c.alignment = center_align
    if end_time:
        c = ws.cell(row=current_row, column=3, value=end_time.strftime('%H:%M'))
        c.alignment = center_align

    prisotnost_mins = 0.0
    if start_time and end_time:
        prisotnost_mins = (end_time - start_time).total_seconds() / 60

    if prisotnost_mins > 0:
        c = ws.cell(row=current_row, column=4, value=td(prisotnost_mins))
        c.number_format = HH_MM_SS
        c.alignment     = center_align

    for col in range(1, 18):
        c = ws.cell(row=current_row, column=col)
        c.border    = thin_border
        c.alignment = center_align

    def write_td(col, mins):
        c = ws.cell(row=current_row, column=col, value=td(mins))
        c.number_format = HH_MM_SS
        c.border        = thin_border
        c.alignment     = center_align
        week_totals[col] += mins
        total_mins[col]  += mins

    write_td(5,  voznja_mins)
    write_td(6,  delo_mins)
    write_td(7,  skupni_mins)
    write_td(8,  0)
    write_td(9,  skupni_mins)
    write_td(10, nocno_delo_mins)
    write_td(11, 0)
    write_td(12, 0)
    write_td(13, razpol_mins)
    write_td(15, rest_mins if rest_mins else (0))

    c = ws.cell(row=current_row, column=17, value='Ne')
    c.border    = thin_border
    c.alignment = center_align


def write_driver_sheet(wb: Workbook, driver_name: str, date_activity: dict,
                       company: str, address: str, city: str,
                       emso: str = '', raw_records: dict = None):
    ws = wb.create_sheet(title=driver_name[:31])

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    def set_row(r):
        ws.row_dimensions[r].height = 30

    font15 = Font(name='Calibri', size=15)
    ws['A1'] = company;          ws['A1'].font = font15;  set_row(1)
    ws['C1'] = driver_name;      ws['C1'].font = font15
    ws['A2'] = address;          ws['A2'].font = font15;  set_row(2)
    ws['C2'] = f'EMŠO: {emso}';  ws['C2'].font = font15
    ws['A3'] = city;             ws['A3'].font = font15;  set_row(3)

    all_dates = all_dates_in_range(list(date_activity.keys()))
    period    = (f'{all_dates[0].strftime("%d.%m.%Y")} - {all_dates[-1].strftime("%d.%m.%Y")}'
                 if all_dates else '')
    ws['A4'] = period; ws['A4'].font = font15; set_row(4)
    set_row(5)

    set_row(6)
    hdr_font     = Font(name='Calibri', size=12, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, hdr in enumerate(HEADERS, start=1):
        c           = ws.cell(row=6, column=col_idx, value=hdr)
        c.font      = hdr_font
        c.border    = thin_border
        c.alignment = center_align

    weeks        = defaultdict(list)
    for d in all_dates:
        weeks[d.isocalendar()[:2]].append(d)

    sorted_weeks = sorted(weeks.keys())
    current_row  = 7
    week_count   = 0
    total_mins   = defaultdict(float)

    for wk_key in sorted_weeks:
        iso_year, iso_week = wk_key
        week_count += 1
        week_dates  = sorted(weeks[wk_key])
        week_totals = defaultdict(float)

        for d in week_dates:
            set_row(current_row)
            acts        = date_activity.get(d, {})
            day_records = []
            if raw_records and driver_name in raw_records and d in raw_records[driver_name]:
                day_records = raw_records[driver_name][d]

            _write_day_row(ws, current_row, d, acts, day_records, week_totals, total_mins)
            current_row += 1

        set_row(current_row)
        wt_cell       = ws.cell(row=current_row, column=1, value=f'Teden {iso_week}')
        wt_cell.fill  = GREY_FILL
        wt_cell.border = thin_border

        for col in range(1, 18):
            c           = ws.cell(row=current_row, column=col)
            c.fill      = GREY_FILL
            c.border    = thin_border
            c.alignment = center_align

        c_i              = ws.cell(row=current_row, column=9, value=td(week_totals[9]))
        c_i.number_format = HH_MM_SS
        c_i.fill         = GREY_FILL
        c_i.border       = thin_border

        c_j              = ws.cell(row=current_row, column=10, value=td(week_totals[10]))
        c_j.number_format = HH_MM_SS
        c_j.fill         = GREY_FILL
        c_j.border       = thin_border

        current_row += 1

    set_row(current_row)
    current_row += 1

    set_row(current_row)
    ws.cell(row=current_row, column=1, value='Seštevek')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
        c               = ws.cell(row=current_row, column=col, value=td(total_mins.get(col, 0)))
        c.number_format = HH_MM_SS
        c.alignment     = center_align

    current_row += 2

    set_row(current_row)
    ws.cell(row=current_row, column=1, value='Mesečna delovna obveznost')
    ws.cell(row=current_row, column=2, value=td(10560))
    ws[f'B{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=4, value='Nabor ur za zadnje 4 mesece *')
    ws.cell(row=current_row, column=5, value=td(32 * 24 * 60))
    ws[f'E{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=13, value='____________________')
    ws.cell(row=current_row, column=15, value='____________________')
    current_row += 1

    set_row(current_row)
    ws.cell(row=current_row, column=1, value='Zmanjšana mesečna delovna obveznost')
    ws.cell(row=current_row, column=2, value=td(10560))
    ws[f'B{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=4, value='Ustvarjene ure v zadnjih 4 mesecih *')
    total_skupni = total_mins.get(9, 0)
    ws.cell(row=current_row, column=5, value=td(total_skupni))
    ws[f'E{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=13, value=driver_name)
    ws.cell(row=current_row, column=15, value=company)
    current_row += 1

    set_row(current_row)
    ws.cell(row=current_row, column=1, value='Opravljene ure')
    ws.cell(row=current_row, column=2, value=td(total_skupni))
    ws[f'B{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=4, value='Tedensko povprečje')
    avg = total_skupni / max(week_count, 1)
    ws.cell(row=current_row, column=5, value=td(avg))
    ws[f'E{current_row}'].number_format = HH_MM_SS
    current_row += 1

    set_row(current_row)
    ws.cell(row=current_row, column=1, value='Presežek')
    ws.cell(row=current_row, column=2, value=td(0))
    ws[f'B{current_row}'].number_format = HH_MM_SS
    ws.cell(row=current_row, column=4, value='* Zadnjih zaključenih 16 tednov')

    return ws


def main():
    if len(sys.argv) < 3:
        print('Uporaba: python json_to_delovni_zapis.py input.json output.xlsx')
        sys.exit(1)

    json_path = sys.argv[1]
    out_path  = sys.argv[2]

    company = "JAKOB d.o.o.o"
    address = "PAKA 4"
    city    = "3205 VITANJE"

    records                             = load_json(json_path)
    date_by_driver, drivers, raw_records = group_by_driver_date(records)

    wb = Workbook()
    wb.remove(wb.active)

    for driver in drivers:
        write_driver_sheet(
            wb, driver, date_by_driver[driver],
            company=company, address=address, city=city,
            raw_records=raw_records
        )

    wb.save(out_path)
    print(f'Shranjeno: {out_path}  ({len(drivers)} list(a): {", ".join(drivers)})')


if __name__ == '__main__':
    main()